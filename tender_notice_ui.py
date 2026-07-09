from __future__ import annotations

import argparse
import json
import threading
import time
import tkinter as tk
import webbrowser
from datetime import datetime, timedelta
from pathlib import Path
from tkinter import filedialog, messagebox, simpledialog, ttk

import requests
import urllib3
from urllib3.exceptions import InsecureRequestWarning

import tender_search


APP_DIR = Path(__file__).resolve().parent
KEYWORDS_FILE = APP_DIR / "keywords.json"


class TenderNoticeApp(ttk.Frame):
    def __init__(self, master: tk.Misc | None = None) -> None:
        super().__init__(master)

        self.rows: list[dict[str, str]] = []
        self.worker: threading.Thread | None = None
        self.stop_requested = False
        self.appeal_rows: list[dict[str, str]] = []
        self.appeal_worker: threading.Thread | None = None
        self.appeal_stop_requested = False
        self.saved_categories = self._load_keyword_categories()
        self.saved_keywords = self._all_saved_keywords()

        self._build_style()
        self._build_ui()

    def _build_style(self) -> None:
        style = ttk.Style(self)
        try:
            style.theme_use("vista")
        except tk.TclError:
            pass
        style.configure("Treeview", rowheight=28)
        style.configure("Primary.TButton", padding=(14, 6))

    def _build_ui(self) -> None:
        self.columnconfigure(0, weight=1)
        self.rowconfigure(1, weight=1)

        title_bar = ttk.Frame(self, padding=(14, 12, 14, 6))
        title_bar.grid(row=0, column=0, sticky="ew")
        title_bar.columnconfigure(0, weight=1)
        ttk.Label(title_bar, text="政府電子採購網標案批次查詢",
                  font=("Microsoft JhengHei UI", 16, "bold")).grid(row=0, column=0, sticky="w")
        self.status_var = tk.StringVar(value="準備就緒")
        ttk.Label(title_bar, textvariable=self.status_var, foreground="#555").grid(row=0, column=1, sticky="e")

        notebook = ttk.Notebook(self)
        notebook.grid(row=1, column=0, sticky="nsew", padx=14, pady=(4, 12))
        self.query_tab = ttk.Frame(notebook, padding=12)
        self.appeal_tab = ttk.Frame(notebook, padding=12)
        self.manage_tab = ttk.Frame(notebook, padding=12)
        notebook.add(self.query_tab, text="招標查詢")
        notebook.add(self.appeal_tab, text="公開徵求查詢")
        notebook.add(self.manage_tab, text="關鍵字管理")

        self.query_tab.columnconfigure(0, weight=1)
        self.query_tab.columnconfigure(1, weight=1)
        self.query_tab.rowconfigure(4, weight=1)

        query_category_box = ttk.Frame(self.query_tab)
        query_category_box.grid(row=0, column=0, sticky="nsew", padx=(0, 8))
        query_category_box.columnconfigure(0, weight=1)
        ttk.Label(query_category_box, text="選分類（可單選或多選；會查分類內全部標案名稱）").grid(row=0, column=0, sticky="w")
        query_category_frame = ttk.Frame(query_category_box)
        query_category_frame.grid(row=1, column=0, sticky="nsew", pady=(4, 0))
        query_category_frame.columnconfigure(0, weight=1)
        self.query_category_listbox = tk.Listbox(
            query_category_frame,
            height=6,
            selectmode=tk.EXTENDED,
            exportselection=False,
            font=("Microsoft JhengHei UI", 10),
        )
        self.query_category_listbox.grid(row=0, column=0, sticky="nsew")
        self.query_category_listbox.bind("<<ListboxSelect>>", lambda _event: self._refresh_query_keyword_listbox())
        self.query_category_listbox.bind(
            "<Button-1>",
            lambda event: self._toggle_category_selection(self.query_category_listbox, event, self._refresh_query_keyword_listbox),
        )
        query_category_scroll = ttk.Scrollbar(query_category_frame, orient="vertical", command=self.query_category_listbox.yview)
        query_category_scroll.grid(row=0, column=1, sticky="ns")
        self.query_category_listbox.configure(yscrollcommand=query_category_scroll.set)
        query_category_actions = ttk.Frame(query_category_box)
        query_category_actions.grid(row=2, column=0, sticky="ew", pady=(6, 0))
        ttk.Button(query_category_actions, text="全選分類", command=self.select_all_categories).pack(side=tk.LEFT)
        ttk.Button(query_category_actions, text="取消分類", command=self.clear_category_selection).pack(side=tk.LEFT, padx=(6, 0))

        query_keyword_box = ttk.Frame(self.query_tab)
        query_keyword_box.grid(row=0, column=1, sticky="nsew", padx=(8, 0))
        query_keyword_box.columnconfigure(0, weight=1)
        ttk.Label(query_keyword_box, text="將查詢的標案名稱預覽").grid(row=0, column=0, sticky="w")
        query_keyword_frame = ttk.Frame(query_keyword_box)
        query_keyword_frame.grid(row=1, column=0, sticky="nsew", pady=(4, 0))
        query_keyword_frame.columnconfigure(0, weight=1)
        self.query_keyword_listbox = tk.Listbox(
            query_keyword_frame,
            height=6,
            selectmode=tk.EXTENDED,
            exportselection=False,
            font=("Microsoft JhengHei UI", 10),
        )
        self.query_keyword_listbox.grid(row=0, column=0, sticky="nsew")
        query_keyword_scroll = ttk.Scrollbar(query_keyword_frame, orient="vertical", command=self.query_keyword_listbox.yview)
        query_keyword_scroll.grid(row=0, column=1, sticky="ns")
        self.query_keyword_listbox.configure(yscrollcommand=query_keyword_scroll.set)
        ttk.Label(
            query_keyword_box,
            text="選取分類才會帶出該分類的標案名稱，會查整個分類；再點一次分類即可取消。也可用下方臨時標案名稱查詢。",
            foreground="#666",
            wraplength=520,
        ).grid(row=2, column=0, sticky="ew", pady=(6, 0))

        manual_box = ttk.Frame(self.query_tab)
        manual_box.grid(row=1, column=0, columnspan=2, sticky="ew", pady=(10, 0))
        manual_box.columnconfigure(0, weight=1)
        ttk.Label(manual_box, text="臨時標案名稱（一行一個，不會存入 JSON）").grid(row=0, column=0, sticky="w")
        self.keyword_text = tk.Text(manual_box, height=3, wrap="word", font=("Microsoft JhengHei UI", 10))
        self.keyword_text.grid(row=1, column=0, sticky="ew", pady=(4, 0))

        self.manage_tab.columnconfigure(0, weight=1)
        self.manage_tab.columnconfigure(1, weight=1)
        self.manage_tab.rowconfigure(0, weight=1)
        manage_category_box = ttk.Frame(self.manage_tab)
        manage_category_box.grid(row=0, column=0, sticky="nsew", padx=(0, 8))
        manage_category_box.columnconfigure(0, weight=1)
        ttk.Label(manage_category_box, text="分類").grid(row=0, column=0, sticky="w")
        manage_category_frame = ttk.Frame(manage_category_box)
        manage_category_frame.grid(row=1, column=0, sticky="nsew", pady=(4, 0))
        manage_category_frame.columnconfigure(0, weight=1)
        self.manage_category_listbox = tk.Listbox(
            manage_category_frame,
            height=6,
            selectmode=tk.BROWSE,
            exportselection=False,
            font=("Microsoft JhengHei UI", 10),
        )
        self.manage_category_listbox.grid(row=0, column=0, sticky="nsew")
        self.manage_category_listbox.bind("<<ListboxSelect>>", lambda _event: self._refresh_manage_keyword_listbox())
        manage_category_scroll = ttk.Scrollbar(manage_category_frame, orient="vertical", command=self.manage_category_listbox.yview)
        manage_category_scroll.grid(row=0, column=1, sticky="ns")
        self.manage_category_listbox.configure(yscrollcommand=manage_category_scroll.set)
        manage_category_actions = ttk.Frame(manage_category_box)
        manage_category_actions.grid(row=2, column=0, sticky="ew", pady=(6, 0))
        ttk.Button(manage_category_actions, text="新增分類", command=self.add_category_dialog).pack(side=tk.LEFT)
        ttk.Button(manage_category_actions, text="刪除選取分類", command=self.delete_selected_categories).pack(side=tk.LEFT, padx=(6, 0))

        manage_keyword_box = ttk.Frame(self.manage_tab)
        manage_keyword_box.grid(row=0, column=1, sticky="nsew", padx=(8, 0))
        manage_keyword_box.columnconfigure(0, weight=1)
        ttk.Label(manage_keyword_box, text="選取分類內的關鍵字").grid(row=0, column=0, sticky="w")
        manage_keyword_frame = ttk.Frame(manage_keyword_box)
        manage_keyword_frame.grid(row=1, column=0, sticky="nsew", pady=(4, 0))
        manage_keyword_frame.columnconfigure(0, weight=1)
        self.manage_keyword_listbox = tk.Listbox(
            manage_keyword_frame,
            height=6,
            selectmode=tk.EXTENDED,
            exportselection=False,
            font=("Microsoft JhengHei UI", 10),
        )
        self.manage_keyword_listbox.grid(row=0, column=0, sticky="nsew")
        manage_keyword_scroll = ttk.Scrollbar(manage_keyword_frame, orient="vertical", command=self.manage_keyword_listbox.yview)
        manage_keyword_scroll.grid(row=0, column=1, sticky="ns")
        self.manage_keyword_listbox.configure(yscrollcommand=manage_keyword_scroll.set)
        manage_keyword_actions = ttk.Frame(manage_keyword_box)
        manage_keyword_actions.grid(row=2, column=0, sticky="ew", pady=(6, 0))
        ttk.Button(manage_keyword_actions, text="新增關鍵字到選取分類", command=self.add_keyword_dialog).pack(side=tk.LEFT)
        ttk.Button(manage_keyword_actions, text="刪除選取關鍵字", command=self.delete_selected_keywords).pack(side=tk.LEFT, padx=(6, 0))
        ttk.Button(manage_keyword_actions, text="匯入關鍵字", command=self.import_keywords).pack(side=tk.LEFT, padx=(6, 0))
        ttk.Button(manage_keyword_actions, text="匯出關鍵字", command=self.export_keywords).pack(side=tk.LEFT, padx=(6, 0))
        ttk.Label(
            self.manage_tab,
            text="管理頁只負責維護 keywords.json；回到「查詢」分頁選分類後即可查詢。",
            foreground="#666",
        ).grid(row=1, column=0, columnspan=2, sticky="w", pady=(8, 0))

        self._refresh_category_listboxes()
        self._refresh_query_keyword_listbox()
        self._refresh_manage_keyword_listbox()

        option_row = ttk.LabelFrame(self.query_tab, text="日期條件", padding=10)
        option_row.grid(row=2, column=0, columnspan=2, sticky="ew", pady=(12, 0))
        self.date_type = tk.StringVar(value="range")
        ttk.Radiobutton(option_row, text="當日", variable=self.date_type, value="today",
                        command=self._toggle_date_entries).pack(side=tk.LEFT)
        ttk.Radiobutton(option_row, text="等標期內", variable=self.date_type, value="period",
                        command=self._toggle_date_entries).pack(side=tk.LEFT, padx=(12, 0))
        ttk.Radiobutton(option_row, text="日期區間", variable=self.date_type, value="range",
                        command=self._toggle_date_entries).pack(side=tk.LEFT, padx=(12, 0))

        ttk.Label(option_row, text="起日").pack(side=tk.LEFT, padx=(22, 4))
        default_start, default_end = self._recent_date_range()
        self.start_date_var = tk.StringVar(value=default_start)
        self.start_date_entry = ttk.Entry(option_row, width=14, textvariable=self.start_date_var)
        self.start_date_entry.pack(side=tk.LEFT)
        ttk.Label(option_row, text="迄日").pack(side=tk.LEFT, padx=(10, 4))
        self.end_date_var = tk.StringVar(value=default_end)
        self.end_date_entry = ttk.Entry(option_row, width=14, textvariable=self.end_date_var)
        self.end_date_entry.pack(side=tk.LEFT)
        ttk.Button(option_row, text="最近 3 天", command=self.apply_recent_dates).pack(side=tk.LEFT, padx=(8, 0))
        ttk.Label(option_row, text="日期請用西元，例如 2026/07/06；民國年也可自動轉換",
                  foreground="#666").pack(side=tk.LEFT, padx=(10, 0))

        button_row = ttk.Frame(self.query_tab)
        button_row.grid(row=3, column=0, columnspan=2, sticky="ew", pady=(12, 8))
        self.search_button = ttk.Button(button_row, text="2. 開始查詢", style="Primary.TButton", command=self.start_search)
        self.search_button.pack(side=tk.LEFT)
        self.stop_button = ttk.Button(button_row, text="停止", command=self.request_stop, state="disabled")
        self.stop_button.pack(side=tk.LEFT, padx=(8, 0))
        ttk.Button(button_row, text="匯出 Excel", command=self.export_excel).pack(side=tk.LEFT, padx=(8, 0))
        ttk.Button(button_row, text="開啟標案頁面", command=self.open_selected).pack(side=tk.LEFT, padx=(8, 0))
        ttk.Button(button_row, text="清除結果", command=self.clear_results).pack(side=tk.LEFT, padx=(8, 0))

        table_frame = ttk.Frame(self.query_tab)
        table_frame.grid(row=4, column=0, columnspan=2, sticky="nsew")
        table_frame.columnconfigure(0, weight=1)
        table_frame.rowconfigure(0, weight=1)

        columns = ("keyword", "agency", "tender_id", "name", "notice", "deadline", "budget", "correction")
        self.tree = ttk.Treeview(table_frame, columns=columns, show="headings", selectmode="browse")
        headings = {
            "keyword": "關鍵字",
            "agency": "機關",
            "tender_id": "案號",
            "name": "標案名稱",
            "notice": "公告日期",
            "deadline": "截止投標",
            "budget": "預算金額",
            "correction": "更正",
        }
        widths = {
            "keyword": 90,
            "agency": 230,
            "tender_id": 140,
            "name": 390,
            "notice": 90,
            "deadline": 90,
            "budget": 110,
            "correction": 52,
        }
        for col in columns:
            self.tree.heading(col, text=headings[col])
            self.tree.column(col, width=widths[col], minwidth=52, stretch=(col == "name"))

        self.tree.grid(row=0, column=0, sticky="nsew")
        self.tree.bind("<Double-1>", lambda _event: self.open_selected())

        y_scroll = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
        y_scroll.grid(row=0, column=1, sticky="ns")
        self.tree.configure(yscrollcommand=y_scroll.set)

        self._toggle_date_entries()
        self._build_appeal_tab()

    def _recent_date_range(self) -> tuple[str, str]:
        today = datetime.now()
        start = today - timedelta(days=2)
        return start.strftime("%Y/%m/%d"), today.strftime("%Y/%m/%d")

    def apply_recent_dates(self) -> None:
        start, end = self._recent_date_range()
        self.date_type.set("range")
        self.start_date_var.set(start)
        self.end_date_var.set(end)
        self._toggle_date_entries()

    def _load_keyword_categories(self) -> dict[str, list[str]]:
        if not KEYWORDS_FILE.exists():
            self._save_categories(
                {
                    "土地地政": ["土地", "地政"],
                    "資訊維護": ["資訊", "維護"],
                }
            )
        try:
            data = json.loads(KEYWORDS_FILE.read_text(encoding="utf-8-sig"))
        except (OSError, json.JSONDecodeError):
            return {}

        categories: dict[str, list[str]] = {}
        if isinstance(data, list):
            categories["未分類"] = self._unique_keywords(str(item) for item in data)
        elif isinstance(data, dict):
            raw_categories = data.get("categories", {})
            if isinstance(raw_categories, dict):
                for name, values in raw_categories.items():
                    if isinstance(values, list):
                        categories[str(name)] = self._unique_keywords(values)
            elif isinstance(raw_categories, list):
                for item in raw_categories:
                    if isinstance(item, dict):
                        name = str(item.get("name", "")).strip()
                        values = item.get("keywords", [])
                        if name and isinstance(values, list):
                            categories[name] = self._unique_keywords(values)

            legacy_keywords = data.get("keywords", [])
            if isinstance(legacy_keywords, list) and legacy_keywords:
                categories.setdefault("未分類", [])
                categories["未分類"] = self._unique_keywords(
                    [*categories["未分類"], *legacy_keywords]
                )

        return {name: values for name, values in categories.items() if name}

    def _save_categories(self, categories: dict[str, list[str]]) -> None:
        data = {
            "categories": {
                name: self._unique_keywords(values)
                for name, values in categories.items()
                if name.strip()
            }
        }
        KEYWORDS_FILE.write_text(
            json.dumps(data, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )

    def _all_saved_keywords(self) -> list[str]:
        keywords = []
        for values in self.saved_categories.values():
            keywords.extend(values)
        return self._unique_keywords(keywords)

    def _unique_keywords(self, values) -> list[str]:
        seen = set()
        keywords = []
        for value in values:
            keyword = str(value).strip()
            if keyword and keyword not in seen:
                seen.add(keyword)
                keywords.append(keyword)
        return keywords

    def _selected_query_categories(self) -> list[str]:
        return [
            self.query_category_listbox.get(index)
            for index in self.query_category_listbox.curselection()
        ]

    def _selected_manage_categories(self) -> list[str]:
        selected = [
            self.manage_category_listbox.get(index)
            for index in self.manage_category_listbox.curselection()
        ]
        return selected[:1]

    def _toggle_category_selection(self, listbox: tk.Listbox, event, refresh) -> str:
        """單擊切換分類選取狀態（可單選或多選），再點一次即可取消。"""
        index = listbox.nearest(event.y)
        bbox = listbox.bbox(index)
        if bbox is None:
            return "break"
        top, height = bbox[1], bbox[3]
        if event.y < top or event.y > top + height:
            return "break"
        if index in listbox.curselection():
            listbox.selection_clear(index)
        else:
            listbox.selection_set(index)
        refresh()
        return "break"

    def _refresh_category_listboxes(self) -> None:
        query_selected = set(getattr(self, "_selected_query_categories", lambda: [])())
        manage_selected = set(getattr(self, "_selected_manage_categories", lambda: [])())
        has_appeal = hasattr(self, "appeal_category_listbox")
        appeal_selected = set(self._selected_appeal_categories()) if has_appeal else set()
        self.query_category_listbox.delete(0, tk.END)
        self.manage_category_listbox.delete(0, tk.END)
        if has_appeal:
            self.appeal_category_listbox.delete(0, tk.END)
        for name in self.saved_categories:
            self.query_category_listbox.insert(tk.END, name)
            self.manage_category_listbox.insert(tk.END, name)
            if has_appeal:
                self.appeal_category_listbox.insert(tk.END, name)
        for index, name in enumerate(self.saved_categories):
            if name in query_selected:
                self.query_category_listbox.selection_set(index)
            if name in manage_selected:
                self.manage_category_listbox.selection_set(index)
            if has_appeal and name in appeal_selected:
                self.appeal_category_listbox.selection_set(index)
        if has_appeal:
            self._refresh_appeal_keyword_listbox()

    def _refresh_query_keyword_listbox(self) -> None:
        self.query_keyword_listbox.delete(0, tk.END)
        # 只有選取分類才帶出該分類的標案名稱；沒選分類就清空
        keywords: list[str] = []
        for category in self._selected_query_categories():
            keywords.extend(self.saved_categories.get(category, []))
        for keyword in self._unique_keywords(keywords):
            self.query_keyword_listbox.insert(tk.END, keyword)

    def _refresh_manage_keyword_listbox(self) -> None:
        self.manage_keyword_listbox.delete(0, tk.END)
        keywords = []
        for category in self._selected_manage_categories():
            keywords.extend(self.saved_categories.get(category, []))
        for keyword in self._unique_keywords(keywords):
            self.manage_keyword_listbox.insert(tk.END, keyword)

    def select_all_categories(self) -> None:
        self.query_category_listbox.selection_set(0, tk.END)
        self._refresh_query_keyword_listbox()

    def clear_category_selection(self) -> None:
        self.query_category_listbox.selection_clear(0, tk.END)
        self.query_keyword_listbox.selection_clear(0, tk.END)
        self._refresh_query_keyword_listbox()

    def add_category_dialog(self) -> None:
        name = simpledialog.askstring("新增分類", "請輸入分類名稱：", parent=self)
        if name is None:
            return
        category = name.strip()
        if not category:
            return
        if category in self.saved_categories:
            messagebox.showinfo("分類已存在", f"「{category}」已經存在。")
            return

        raw = simpledialog.askstring(
            "分類關鍵字",
            "請輸入此分類的關鍵字，可用逗號或空白分隔：",
            parent=self,
        )
        keywords = self._split_keywords(raw or "")
        self.saved_categories[category] = keywords
        self.saved_keywords = self._all_saved_keywords()
        self._save_categories(self.saved_categories)
        self._refresh_category_listboxes()
        self.manage_category_listbox.selection_clear(0, tk.END)
        index = list(self.saved_categories).index(category)
        self.manage_category_listbox.selection_set(index)
        self._refresh_manage_keyword_listbox()
        self._refresh_query_keyword_listbox()
        self.status_var.set(f"已新增分類：{category}")

    def delete_selected_categories(self) -> None:
        names = self._selected_manage_categories()
        if not names:
            messagebox.showinfo("尚未選取", "請先選取要刪除的分類。")
            return
        if not messagebox.askyesno("刪除分類", f"確定刪除分類：{'、'.join(names)}？"):
            return
        for name in names:
            self.saved_categories.pop(name, None)
        self.saved_keywords = self._all_saved_keywords()
        self._save_categories(self.saved_categories)
        self._refresh_category_listboxes()
        self._refresh_query_keyword_listbox()
        self._refresh_manage_keyword_listbox()
        self.status_var.set("已更新 keywords.json")

    def add_keyword_dialog(self) -> None:
        selected_categories = self._selected_manage_categories()
        if not selected_categories:
            messagebox.showinfo("請先選分類", "請先選取要加入關鍵字的分類。")
            return
        target_category = selected_categories[0]
        value = simpledialog.askstring(
            "新增關鍵字",
            f"將關鍵字加入分類：{target_category}\n\n"
            "請輸入關鍵字，可用逗號、頓號、空白或換行分隔：",
            parent=self,
        )
        if value is None:
            return
        keywords = self._split_keywords(value)
        if not keywords:
            return
        current = self.saved_categories.setdefault(target_category, [])
        self.saved_categories[target_category] = self._unique_keywords([*current, *keywords])
        self.saved_keywords = self._all_saved_keywords()
        self._save_categories(self.saved_categories)
        self._refresh_query_keyword_listbox()
        self._refresh_manage_keyword_listbox()
        self._refresh_appeal_keyword_listbox()
        self.status_var.set(f"已加入關鍵字：{'、'.join(keywords)}")

    def delete_selected_keywords(self) -> None:
        selected_categories = self._selected_manage_categories()
        if not selected_categories:
            messagebox.showinfo("請先選分類", "請先選取分類，再刪除該分類內的關鍵字。")
            return
        target_category = selected_categories[0]
        selected = set(self.manage_keyword_listbox.curselection())
        if not selected:
            messagebox.showinfo("尚未選取", "請先選取要刪除的關鍵字。")
            return
        names = [self.manage_keyword_listbox.get(index) for index in sorted(selected)]
        if not messagebox.askyesno("刪除關鍵字", f"確定從「{target_category}」刪除：{'、'.join(names)}？"):
            return
        self.saved_categories[target_category] = [
            keyword for keyword in self.saved_categories.get(target_category, [])
            if keyword not in names
        ]
        self.saved_keywords = self._all_saved_keywords()
        self._save_categories(self.saved_categories)
        self._refresh_query_keyword_listbox()
        self._refresh_manage_keyword_listbox()
        self._refresh_appeal_keyword_listbox()
        self.status_var.set("已更新 keywords.json")

    def import_keywords(self) -> None:
        path = filedialog.askopenfilename(
            title="匯入關鍵字 JSON",
            filetypes=[("JSON 檔案", "*.json"), ("所有檔案", "*.*")],
        )
        if not path:
            return
        try:
            categories = self._read_categories_file(Path(path))
        except ValueError as exc:
            messagebox.showerror("匯入失敗", str(exc))
            return
        if not categories:
            messagebox.showwarning("沒有資料", "選取的 JSON 沒有可匯入的分類或關鍵字。")
            return
        if not messagebox.askyesno(
            "匯入關鍵字",
            "匯入後會覆蓋目前的 keywords.json。\n\n確定要匯入嗎？",
        ):
            return
        self.saved_categories = categories
        self.saved_keywords = self._all_saved_keywords()
        self._save_categories(self.saved_categories)
        self._refresh_category_listboxes()
        self._refresh_query_keyword_listbox()
        self._refresh_manage_keyword_listbox()
        self.status_var.set(f"已匯入：{path}")

    def export_keywords(self) -> None:
        default_name = f"關鍵字分類_{datetime.now():%Y%m%d_%H%M%S}.json"
        path = filedialog.asksaveasfilename(
            title="匯出關鍵字 JSON",
            defaultextension=".json",
            initialfile=default_name,
            filetypes=[("JSON 檔案", "*.json"), ("所有檔案", "*.*")],
        )
        if not path:
            return
        data = {
            "categories": {
                name: self._unique_keywords(values)
                for name, values in self.saved_categories.items()
            }
        }
        Path(path).write_text(
            json.dumps(data, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        self.status_var.set(f"已匯出：{path}")
        messagebox.showinfo("匯出完成", "關鍵字分類已匯出。")

    def _read_categories_file(self, path: Path) -> dict[str, list[str]]:
        try:
            data = json.loads(path.read_text(encoding="utf-8-sig"))
        except (OSError, json.JSONDecodeError) as exc:
            raise ValueError(f"無法讀取 JSON：{exc}") from exc

        categories: dict[str, list[str]] = {}
        if isinstance(data, list):
            categories["未分類"] = self._unique_keywords(data)
        elif isinstance(data, dict):
            raw_categories = data.get("categories", {})
            if isinstance(raw_categories, dict):
                for name, values in raw_categories.items():
                    if isinstance(values, list):
                        categories[str(name).strip()] = self._unique_keywords(values)
            elif isinstance(raw_categories, list):
                for item in raw_categories:
                    if isinstance(item, dict):
                        name = str(item.get("name", "")).strip()
                        values = item.get("keywords", [])
                        if isinstance(values, list):
                            categories[name] = self._unique_keywords(values)

            legacy_keywords = data.get("keywords", [])
            if isinstance(legacy_keywords, list):
                categories["未分類"] = self._unique_keywords(legacy_keywords)
        else:
            raise ValueError("JSON 格式不支援，請使用 categories 或 keywords 格式。")

        return {
            name: values
            for name, values in categories.items()
            if name and values
        }

    def _split_keywords(self, value: str) -> list[str]:
        normalized = value.replace("，", ",").replace("、", ",").replace("\n", ",")
        parts = []
        for chunk in normalized.split(","):
            parts.extend(chunk.split())
        return self._unique_keywords(parts)

    def _toggle_date_entries(self) -> None:
        enabled = self.date_type.get() == "range"
        state = "normal" if enabled else "disabled"
        self.start_date_entry.configure(state=state)
        self.end_date_entry.configure(state=state)

    def _keywords(self) -> list[str]:
        category_keywords = []
        selected_categories = self._selected_query_categories()
        for category in selected_categories:
            category_keywords.extend(self.saved_categories.get(category, []))
        selected = []
        if not selected_categories:
            selected = [
                self.query_keyword_listbox.get(index)
                for index in self.query_keyword_listbox.curselection()
            ]
        manual = []
        for line in self.keyword_text.get("1.0", "end").splitlines():
            manual.append(line.strip())
        return self._unique_keywords([*category_keywords, *selected, *manual])

    def _make_args(self) -> argparse.Namespace:
        return argparse.Namespace(
            page_size=100,
            firstSearch="true",
            searchType="basic",
            isBinding="N",
            isLogIn="N",
            org_name="",
            org_id="",
            tender_id="",
            tender_type="TENDER_DECLARATION",
            tender_way="TENDER_WAY_ALL_DECLARATION",
            date_type=self.date_type.get(),
            start_date=self.start_date_var.get().strip(),
            end_date=self.end_date_var.get().strip(),
            category="",
            policy_advocacy="",
            delay=0.45,
            timeout=30,
            verify_ssl=False,
            no_dedupe=False,
        )

    def start_search(self) -> None:
        if self.worker and self.worker.is_alive():
            return

        keywords = self._keywords()
        if not keywords:
            messagebox.showwarning("缺少關鍵字", "請至少輸入一個標案名稱關鍵字。")
            return

        args = self._make_args()
        if args.date_type == "range":
            try:
                tender_search.normalize_roc_date(args.start_date)
                tender_search.normalize_roc_date(args.end_date)
            except ValueError as exc:
                messagebox.showwarning("日期格式錯誤", str(exc))
                return

        self.clear_results()
        self.stop_requested = False
        self.search_button.configure(state="disabled")
        self.stop_button.configure(state="normal")
        self.status_var.set("查詢中...")

        self.worker = threading.Thread(target=self._search_worker, args=(keywords, args), daemon=True)
        self.worker.start()

    def request_stop(self) -> None:
        self.stop_requested = True
        self.status_var.set("停止中，等待目前查詢完成...")

    def _search_worker(self, keywords: list[str], args: argparse.Namespace) -> None:
        try:
            session = requests.Session()
            session.verify = args.verify_ssl
            urllib3.disable_warnings(InsecureRequestWarning)
            session.headers.update(
                {
                    "User-Agent": (
                        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126 Safari/537.36"
                    )
                }
            )

            rows: list[dict[str, str]] = []
            for index, keyword in enumerate(keywords, start=1):
                if self.stop_requested:
                    break
                self._set_status(f"查詢 {index}/{len(keywords)}：{keyword}")
                rows.extend(tender_search.search_keyword(session, keyword, args))
                if index < len(keywords):
                    time.sleep(args.delay)

            if not args.no_dedupe:
                rows = tender_search.dedupe_rows(rows)
            self.after(0, self._finish_search, rows, None, self.stop_requested)
        except Exception as exc:
            self.after(0, self._finish_search, [], exc, False)

    def _set_status(self, text: str) -> None:
        self.after(0, lambda: self.status_var.set(text))

    def _finish_search(self, rows: list[dict[str, str]], error: Exception | None, stopped: bool) -> None:
        self.search_button.configure(state="normal")
        self.stop_button.configure(state="disabled")

        if error:
            self.status_var.set("查詢失敗")
            messagebox.showerror("查詢失敗", str(error))
            return

        self.rows = rows
        for row in rows:
            self.tree.insert(
                "",
                "end",
                values=(
                    row["query_keyword"],
                    row["agency"],
                    row["tender_id"],
                    row["tender_name"],
                    row["announcement_date"],
                    row["bid_deadline"],
                    row["budget_amount"],
                    "是" if row["is_correction"] == "Y" else "",
                ),
            )
        prefix = "已停止，" if stopped else "完成，"
        self.status_var.set(f"{prefix}共 {len(rows)} 筆")
        if not rows and not stopped:
            messagebox.showinfo(
                "查無資料",
                "目前條件沒有查到資料。\n\n"
                "建議改用「等標期內」，或使用「日期區間」放寬公告日期後再查一次。",
            )

    def clear_results(self) -> None:
        self.rows = []
        for item in self.tree.get_children():
            self.tree.delete(item)
        self.status_var.set("準備就緒")

    def _selected_row(self) -> dict[str, str] | None:
        selection = self.tree.selection()
        if not selection:
            return None
        index = self.tree.index(selection[0])
        if 0 <= index < len(self.rows):
            return self.rows[index]
        return None

    def open_selected(self) -> None:
        row = self._selected_row()
        if not row:
            messagebox.showinfo("尚未選取", "請先選取一筆標案。")
            return
        if not row["detail_url"]:
            messagebox.showinfo("沒有連結", "這筆資料沒有可開啟的標案連結。")
            return
        webbrowser.open(row["detail_url"])

    def export_excel(self) -> None:
        if not self.rows:
            messagebox.showinfo("沒有資料", "目前沒有可匯出的查詢結果。")
            return

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError:
            messagebox.showerror(
                "缺少套件",
                "匯出 Excel 需要 openpyxl。\n請重新執行 run_ui.bat，或執行：py -m pip install -r requirements.txt",
            )
            return

        default_name = f"標案查詢結果_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
        path = filedialog.asksaveasfilename(
            title="匯出 Excel",
            defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=[("Excel 檔案", "*.xlsx"), ("所有檔案", "*.*")],
        )
        if not path:
            return

        headers = [
            ("query_keyword", "查詢關鍵字"),
            ("agency", "機關名稱"),
            ("tender_id", "標案案號"),
            ("is_correction", "更正公告"),
            ("tender_name", "標案名稱"),
            ("transmission_count", "傳輸次數"),
            ("tender_method", "招標方式"),
            ("procurement_category", "採購性質"),
            ("announcement_date", "公告日期"),
            ("bid_deadline", "截止投標"),
            ("budget_amount", "預算金額"),
            ("detail_url", "標案連結"),
        ]

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "查詢結果"

        header_fill = PatternFill("solid", fgColor="D9EAF7")
        header_font = Font(bold=True, color="000000")
        for col_idx, (_key, label) in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col_idx, value=label)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row_idx, row in enumerate(self.rows, start=2):
            for col_idx, (key, _label) in enumerate(headers, start=1):
                value = row.get(key, "")
                cell = sheet.cell(row=row_idx, column=col_idx)
                if key == "is_correction":
                    cell.value = "是" if value == "Y" else ""
                elif key == "budget_amount":
                    cell.value = self._parse_budget(value)
                    if cell.value is not None:
                        cell.number_format = '#,##0'
                    else:
                        cell.value = value
                elif key == "detail_url":
                    cell.value = "開啟連結" if value else ""
                    if value:
                        cell.hyperlink = value
                        cell.style = "Hyperlink"
                else:
                    cell.value = value
                cell.alignment = Alignment(vertical="top", wrap_text=(key == "tender_name"))

        widths = {
            "A": 16,
            "B": 28,
            "C": 18,
            "D": 10,
            "E": 52,
            "F": 10,
            "G": 22,
            "H": 12,
            "I": 12,
            "J": 12,
            "K": 14,
            "L": 14,
        }
        for column, width in widths.items():
            sheet.column_dimensions[column].width = width

        sheet.freeze_panes = "A2"
        sheet.row_dimensions[1].height = 24
        for row_idx in range(2, len(self.rows) + 2):
            sheet.row_dimensions[row_idx].height = 36

        workbook.save(path)
        self.status_var.set(f"已匯出：{path}")
        messagebox.showinfo("匯出完成", f"已匯出 {len(self.rows)} 筆資料到 Excel。")

    def _parse_budget(self, value: str):
        text = str(value).replace(",", "").strip()
        if not text:
            return None
        try:
            return int(text)
        except ValueError:
            return None

    # ---------- 公開徵求查詢 ----------

    def _build_appeal_tab(self) -> None:
        self.appeal_tab.columnconfigure(0, weight=1)
        self.appeal_tab.columnconfigure(1, weight=1)
        self.appeal_tab.rowconfigure(4, weight=1)

        category_box = ttk.Frame(self.appeal_tab)
        category_box.grid(row=0, column=0, sticky="nsew", padx=(0, 8))
        category_box.columnconfigure(0, weight=1)
        ttk.Label(category_box, text="選分類（可單選或多選；會查分類內全部標案名稱）").grid(row=0, column=0, sticky="w")
        category_frame = ttk.Frame(category_box)
        category_frame.grid(row=1, column=0, sticky="nsew", pady=(4, 0))
        category_frame.columnconfigure(0, weight=1)
        self.appeal_category_listbox = tk.Listbox(
            category_frame,
            height=6,
            selectmode=tk.EXTENDED,
            exportselection=False,
            font=("Microsoft JhengHei UI", 10),
        )
        self.appeal_category_listbox.grid(row=0, column=0, sticky="nsew")
        self.appeal_category_listbox.bind("<<ListboxSelect>>", lambda _event: self._refresh_appeal_keyword_listbox())
        self.appeal_category_listbox.bind(
            "<Button-1>",
            lambda event: self._toggle_category_selection(self.appeal_category_listbox, event, self._refresh_appeal_keyword_listbox),
        )
        category_scroll = ttk.Scrollbar(category_frame, orient="vertical", command=self.appeal_category_listbox.yview)
        category_scroll.grid(row=0, column=1, sticky="ns")
        self.appeal_category_listbox.configure(yscrollcommand=category_scroll.set)
        category_actions = ttk.Frame(category_box)
        category_actions.grid(row=2, column=0, sticky="ew", pady=(6, 0))
        ttk.Button(category_actions, text="全選分類", command=self.select_all_appeal_categories).pack(side=tk.LEFT)
        ttk.Button(category_actions, text="取消分類", command=self.clear_appeal_category_selection).pack(side=tk.LEFT, padx=(6, 0))

        keyword_box = ttk.Frame(self.appeal_tab)
        keyword_box.grid(row=0, column=1, sticky="nsew", padx=(8, 0))
        keyword_box.columnconfigure(0, weight=1)
        ttk.Label(keyword_box, text="將查詢的標案名稱預覽").grid(row=0, column=0, sticky="w")
        keyword_frame = ttk.Frame(keyword_box)
        keyword_frame.grid(row=1, column=0, sticky="nsew", pady=(4, 0))
        keyword_frame.columnconfigure(0, weight=1)
        self.appeal_keyword_listbox = tk.Listbox(
            keyword_frame,
            height=6,
            selectmode=tk.EXTENDED,
            exportselection=False,
            font=("Microsoft JhengHei UI", 10),
        )
        self.appeal_keyword_listbox.grid(row=0, column=0, sticky="nsew")
        keyword_scroll = ttk.Scrollbar(keyword_frame, orient="vertical", command=self.appeal_keyword_listbox.yview)
        keyword_scroll.grid(row=0, column=1, sticky="ns")
        self.appeal_keyword_listbox.configure(yscrollcommand=keyword_scroll.set)
        ttk.Label(
            keyword_box,
            text="選取分類才會帶出該分類的標案名稱；取消分類即清空。也可用下方臨時標案名稱查詢。",
            foreground="#666",
            wraplength=520,
        ).grid(row=2, column=0, sticky="ew", pady=(6, 0))

        manual_box = ttk.Frame(self.appeal_tab)
        manual_box.grid(row=1, column=0, columnspan=2, sticky="ew", pady=(10, 0))
        manual_box.columnconfigure(0, weight=1)
        ttk.Label(manual_box, text="臨時標案名稱（一行一個，不會存入 JSON）").grid(row=0, column=0, sticky="w")
        self.appeal_keyword_text = tk.Text(manual_box, height=3, wrap="word", font=("Microsoft JhengHei UI", 10))
        self.appeal_keyword_text.grid(row=1, column=0, sticky="ew", pady=(4, 0))

        date_row = ttk.LabelFrame(self.appeal_tab, text="公開徵求日期（起訖）", padding=10)
        date_row.grid(row=2, column=0, columnspan=2, sticky="ew", pady=(12, 0))
        default_start, default_end = self._recent_appeal_range()
        ttk.Label(date_row, text="起日").pack(side=tk.LEFT, padx=(0, 4))
        self.appeal_start_date_var = tk.StringVar(value=default_start)
        ttk.Entry(date_row, width=14, textvariable=self.appeal_start_date_var).pack(side=tk.LEFT)
        ttk.Label(date_row, text="迄日").pack(side=tk.LEFT, padx=(10, 4))
        self.appeal_end_date_var = tk.StringVar(value=default_end)
        ttk.Entry(date_row, width=14, textvariable=self.appeal_end_date_var).pack(side=tk.LEFT)
        ttk.Button(date_row, text="最近 3 天", command=self.apply_recent_appeal_dates).pack(side=tk.LEFT, padx=(8, 0))
        ttk.Label(date_row, text="日期用西元，例如 2026/07/09；民國年也可；區間請勿超過 3 個月",
                  foreground="#666").pack(side=tk.LEFT, padx=(10, 0))

        button_row = ttk.Frame(self.appeal_tab)
        button_row.grid(row=3, column=0, columnspan=2, sticky="ew", pady=(12, 8))
        self.appeal_search_button = ttk.Button(button_row, text="開始查詢", style="Primary.TButton", command=self.start_appeal_search)
        self.appeal_search_button.pack(side=tk.LEFT)
        self.appeal_stop_button = ttk.Button(button_row, text="停止", command=self.request_appeal_stop, state="disabled")
        self.appeal_stop_button.pack(side=tk.LEFT, padx=(8, 0))
        ttk.Button(button_row, text="匯出 Excel", command=self.appeal_export_excel).pack(side=tk.LEFT, padx=(8, 0))
        ttk.Button(button_row, text="開啟徵求頁面", command=self.appeal_open_selected).pack(side=tk.LEFT, padx=(8, 0))
        ttk.Button(button_row, text="清除結果", command=self.appeal_clear_results).pack(side=tk.LEFT, padx=(8, 0))

        table_frame = ttk.Frame(self.appeal_tab)
        table_frame.grid(row=4, column=0, columnspan=2, sticky="nsew")
        table_frame.columnconfigure(0, weight=1)
        table_frame.rowconfigure(0, weight=1)

        columns = ("keyword", "agency", "tender_id", "name", "appeal_date", "count")
        self.appeal_tree = ttk.Treeview(table_frame, columns=columns, show="headings", selectmode="browse")
        headings = {
            "keyword": "查詢名稱",
            "agency": "機關",
            "tender_id": "案號",
            "name": "標案名稱",
            "appeal_date": "公開徵求日期",
            "count": "公告次數",
        }
        widths = {
            "keyword": 100,
            "agency": 240,
            "tender_id": 150,
            "name": 400,
            "appeal_date": 190,
            "count": 70,
        }
        for col in columns:
            self.appeal_tree.heading(col, text=headings[col])
            self.appeal_tree.column(col, width=widths[col], minwidth=52, stretch=(col == "name"))
        self.appeal_tree.grid(row=0, column=0, sticky="nsew")
        self.appeal_tree.bind("<Double-1>", lambda _event: self.appeal_open_selected())
        appeal_y_scroll = ttk.Scrollbar(table_frame, orient="vertical", command=self.appeal_tree.yview)
        appeal_y_scroll.grid(row=0, column=1, sticky="ns")
        self.appeal_tree.configure(yscrollcommand=appeal_y_scroll.set)

        self._refresh_category_listboxes()

    def _recent_appeal_range(self) -> tuple[str, str]:
        today = datetime.now()
        start = today - timedelta(days=2)
        return start.strftime("%Y/%m/%d"), today.strftime("%Y/%m/%d")

    def apply_recent_appeal_dates(self) -> None:
        start, end = self._recent_appeal_range()
        self.appeal_start_date_var.set(start)
        self.appeal_end_date_var.set(end)

    def _selected_appeal_categories(self) -> list[str]:
        return [
            self.appeal_category_listbox.get(index)
            for index in self.appeal_category_listbox.curselection()
        ]

    def _refresh_appeal_keyword_listbox(self) -> None:
        if not hasattr(self, "appeal_keyword_listbox"):
            return
        self.appeal_keyword_listbox.delete(0, tk.END)
        # 只有選取分類才帶出該分類的標案名稱；沒選分類就清空
        keywords: list[str] = []
        for category in self._selected_appeal_categories():
            keywords.extend(self.saved_categories.get(category, []))
        for keyword in self._unique_keywords(keywords):
            self.appeal_keyword_listbox.insert(tk.END, keyword)

    def select_all_appeal_categories(self) -> None:
        self.appeal_category_listbox.selection_set(0, tk.END)
        self._refresh_appeal_keyword_listbox()

    def clear_appeal_category_selection(self) -> None:
        self.appeal_category_listbox.selection_clear(0, tk.END)
        self.appeal_keyword_listbox.selection_clear(0, tk.END)
        self._refresh_appeal_keyword_listbox()

    def _appeal_keywords(self) -> list[str]:
        category_keywords = []
        selected_categories = self._selected_appeal_categories()
        for category in selected_categories:
            category_keywords.extend(self.saved_categories.get(category, []))
        selected = []
        if not selected_categories:
            selected = [
                self.appeal_keyword_listbox.get(index)
                for index in self.appeal_keyword_listbox.curselection()
            ]
        manual = [line.strip() for line in self.appeal_keyword_text.get("1.0", "end").splitlines()]
        return self._unique_keywords([*category_keywords, *selected, *manual])

    def start_appeal_search(self) -> None:
        if self.appeal_worker and self.appeal_worker.is_alive():
            return

        start_date = self.appeal_start_date_var.get().strip()
        end_date = self.appeal_end_date_var.get().strip()
        if not start_date or not end_date:
            messagebox.showwarning("缺少日期", "請輸入公開徵求日期的起日與迄日。")
            return
        try:
            start_norm = tender_search.normalize_roc_date(start_date)
            end_norm = tender_search.normalize_roc_date(end_date)
        except ValueError as exc:
            messagebox.showwarning("日期格式錯誤", str(exc))
            return
        start_dt = datetime.strptime(start_norm, "%Y/%m/%d")
        end_dt = datetime.strptime(end_norm, "%Y/%m/%d")
        if end_dt < start_dt:
            messagebox.showwarning("日期顛倒", "迄日不可早於起日。")
            return
        if (end_dt - start_dt).days > 93:
            messagebox.showwarning("區間過長", "公開徵求查詢的日期區間請勿超過 3 個月，請縮短後再查。")
            return

        keywords = self._appeal_keywords()
        if not keywords:
            messagebox.showwarning("缺少查詢條件", "請先選取分類，或輸入臨時標案名稱後再查詢。")
            return

        self.appeal_clear_results()
        self.appeal_stop_requested = False
        self.appeal_search_button.configure(state="disabled")
        self.appeal_stop_button.configure(state="normal")
        self.status_var.set("公開徵求查詢中...")

        self.appeal_worker = threading.Thread(
            target=self._appeal_search_worker,
            args=(keywords, start_norm, end_norm),
            daemon=True,
        )
        self.appeal_worker.start()

    def request_appeal_stop(self) -> None:
        self.appeal_stop_requested = True
        self.status_var.set("停止中，等待目前查詢完成...")

    def _appeal_search_worker(self, keywords: list[str], start_date: str, end_date: str) -> None:
        try:
            session = requests.Session()
            session.verify = False
            urllib3.disable_warnings(InsecureRequestWarning)
            session.headers.update(
                {
                    "User-Agent": (
                        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126 Safari/537.36"
                    )
                }
            )

            rows: list[dict[str, str]] = []
            for index, keyword in enumerate(keywords, start=1):
                if self.appeal_stop_requested:
                    break
                label = keyword or "(全部)"
                self._set_status(f"公開徵求查詢 {index}/{len(keywords)}：{label}")
                rows.extend(
                    tender_search.search_appeal(session, keyword, start_date, end_date, delay=0.45, timeout=30)
                )
                if index < len(keywords):
                    time.sleep(0.45)

            rows = tender_search.dedupe_appeal_rows(rows)
            self.after(0, self._finish_appeal_search, rows, None, self.appeal_stop_requested)
        except Exception as exc:
            self.after(0, self._finish_appeal_search, [], exc, False)

    def _finish_appeal_search(self, rows: list[dict[str, str]], error: Exception | None, stopped: bool) -> None:
        self.appeal_search_button.configure(state="normal")
        self.appeal_stop_button.configure(state="disabled")

        if error:
            self.status_var.set("公開徵求查詢失敗")
            messagebox.showerror("查詢失敗", str(error))
            return

        self.appeal_rows = rows
        for row in rows:
            self.appeal_tree.insert(
                "",
                "end",
                values=(
                    row["query_keyword"],
                    row["agency"],
                    row["tender_id"],
                    row["tender_name"],
                    self._appeal_date_text(row),
                    row["announcement_count"],
                ),
            )
        prefix = "已停止，" if stopped else "完成，"
        self.status_var.set(f"{prefix}公開徵求共 {len(rows)} 筆")
        if not rows and not stopped:
            messagebox.showinfo("查無資料", "目前日期區間沒有查到公開徵求資料，請調整標案名稱或日期後再查一次。")

    def _appeal_date_text(self, row: dict[str, str]) -> str:
        announce = row.get("announcement_date", "")
        deadline = row.get("deadline", "")
        if announce and deadline:
            return f"{announce} ~ {deadline}"
        return announce or deadline

    def appeal_clear_results(self) -> None:
        self.appeal_rows = []
        for item in self.appeal_tree.get_children():
            self.appeal_tree.delete(item)
        self.status_var.set("準備就緒")

    def _selected_appeal_row(self) -> dict[str, str] | None:
        selection = self.appeal_tree.selection()
        if not selection:
            return None
        index = self.appeal_tree.index(selection[0])
        if 0 <= index < len(self.appeal_rows):
            return self.appeal_rows[index]
        return None

    def appeal_open_selected(self) -> None:
        row = self._selected_appeal_row()
        if not row:
            messagebox.showinfo("尚未選取", "請先選取一筆公開徵求。")
            return
        if not row["detail_url"]:
            messagebox.showinfo("沒有連結", "這筆資料沒有可開啟的徵求連結。")
            return
        webbrowser.open(row["detail_url"])

    def appeal_export_excel(self) -> None:
        if not self.appeal_rows:
            messagebox.showinfo("沒有資料", "目前沒有可匯出的公開徵求查詢結果。")
            return

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill
        except ImportError:
            messagebox.showerror(
                "缺少套件",
                "匯出 Excel 需要 openpyxl。\n請重新執行 run_ui.bat，或執行：py -m pip install -r requirements.txt",
            )
            return

        default_name = f"公開徵求查詢結果_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
        path = filedialog.asksaveasfilename(
            title="匯出 Excel",
            defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=[("Excel 檔案", "*.xlsx"), ("所有檔案", "*.*")],
        )
        if not path:
            return

        headers = [
            ("query_keyword", "查詢名稱"),
            ("agency", "機關名稱"),
            ("tender_id", "標案案號"),
            ("tender_name", "標案名稱"),
            ("announcement_count", "公告次數"),
            ("appeal_date", "公開徵求日期"),
            ("detail_url", "徵求連結"),
        ]

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "公開徵求"

        header_fill = PatternFill("solid", fgColor="D9EAF7")
        header_font = Font(bold=True, color="000000")
        for col_idx, (_key, label) in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col_idx, value=label)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row_idx, row in enumerate(self.appeal_rows, start=2):
            for col_idx, (key, _label) in enumerate(headers, start=1):
                value = row.get(key, "")
                cell = sheet.cell(row=row_idx, column=col_idx)
                if key == "appeal_date":
                    cell.value = self._appeal_date_text(row)
                elif key == "detail_url":
                    cell.value = "開啟連結" if value else ""
                    if value:
                        cell.hyperlink = value
                        cell.style = "Hyperlink"
                else:
                    cell.value = value
                cell.alignment = Alignment(vertical="top", wrap_text=(key == "tender_name"))

        widths = {"A": 16, "B": 30, "C": 20, "D": 52, "E": 10, "F": 22, "G": 14}
        for column, width in widths.items():
            sheet.column_dimensions[column].width = width

        sheet.freeze_panes = "A2"
        sheet.row_dimensions[1].height = 24
        for row_idx in range(2, len(self.appeal_rows) + 2):
            sheet.row_dimensions[row_idx].height = 36

        workbook.save(path)
        self.status_var.set(f"已匯出：{path}")
        messagebox.showinfo("匯出完成", f"已匯出 {len(self.appeal_rows)} 筆公開徵求資料到 Excel。")


def main() -> None:
    root = tk.Tk()
    root.title("標案批次查詢")
    root.geometry("1220x760")
    root.minsize(960, 560)
    app = TenderNoticeApp(root)
    app.pack(fill="both", expand=True)
    root.mainloop()


if __name__ == "__main__":
    main()
